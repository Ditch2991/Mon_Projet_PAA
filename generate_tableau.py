"""
generate_tableau.py — Export Excel PAA style officiel
======================================================
Reproduit fidelement le format du tableau de reference PAA :
  En-tete vert clair (E2EFDA), Totaux bleu vif (00B0F0),
  Groupes colores par niveau, Colonne hypothese a droite,
  Section Escales + Marchandises dans un seul tableau unifie.
"""

import io
import pickle
import numpy as np
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

# ─── PALETTE (fidelite tableau PAA reference) ─────────────────────
C_ENTETE  = "E2EFDA"  # vert clair  : titre, annees, realise label
C_TOTAL   = "00B0F0"  # bleu vif    : totaux principaux
C_GRP1    = "156082"  # bleu PAA    : Import/Export, National, Transb
C_GRP2    = "E97132"  # orange      : Port de Commerce, groupes niv.2
C_GRP3    = "A02B93"  # violet      : Prod petroliers, March. gen.
C_BLANC   = "FFFFFF"  # blanc       : lignes detail
C_GRIS    = "F2F2F2"  # gris clair  : croissance, IC
C_REEL_FG = "FF0000"  # rouge       : texte valeur realisee
C_PREV_FG = "000000"  # noir        : texte valeurs prevues
C_TITRE   = "000000"  # noir        : texte en-tete
C_GRP1_FG = "FFFFFF"  # blanc       : texte sur bleu PAA
C_GRP2_FG = "FFFFFF"  # blanc       : texte sur orange
C_GRP3_FG = "FFFFFF"  # blanc       : texte sur violet

MOIS_NOMS = ["Janvier","F\u00e9vrier","Mars","Avril","Mai","Juin",
             "Juillet","Ao\u00fbt","Septembre","Octobre","Novembre","D\u00e9cembre"]

SEGS_ESC = ["TC1","TC2","C\u00e9r\u00e9alier","Fruitier","Min\u00e9ralier",
            "P\u00e9trolier","Roulier","Quai Nord","Quai Ouest","Autres zones"]

SEG_TO_AXE = {
    "Import": "sens", "Export": "sens",
    "March. g\u00e9n\u00e9rales": "composante",
    "Prod. p\u00e9troliers": "composante",
    "Prod. de p\u00eache": "composante",
    "National": "destination", "Transit": "destination",
    "Transbordement": "destination",
    "Conteneuris\u00e9": "conteneur",
    "Non conteneuris\u00e9": "conteneur",
}

AXE_TO_SEGS = {
    "sens"       : ["Import", "Export"],
    "composante" : ["March. g\u00e9n\u00e9rales", "Prod. p\u00e9troliers",
                    "Prod. de p\u00eache"],
    "destination": ["National", "Transit", "Transbordement"],
    "conteneur"  : ["Conteneuris\u00e9", "Non conteneuris\u00e9"],
}

APPROCHE_LABELS = {
    "top_down"       : "Du global vers le d\u00e9tail (Top-down)",
    "bu_sens"        : "Du d\u00e9tail vers le global \u2014 Sens",
    "bu_composante"  : "Du d\u00e9tail vers le global \u2014 Composante",
    "bu_destination" : "Du d\u00e9tail vers le global \u2014 Destination",
    "bu_conteneur"   : "Du d\u00e9tail vers le global \u2014 Conteneurisation",
}

# ─── HELPERS STYLE ────────────────────────────────────────────────
def _fill(h): return PatternFill("solid", fgColor=h)
def _font(bold=False, color="000000", size=10, italic=False):
    return Font(name="Arial", bold=bold, color=color, size=size, italic=italic)
def _align(h="center", wrap=True):
    return Alignment(horizontal=h, vertical="center", wrap_text=wrap)
def _border():
    s = Side(style="thin", color="CCCCCC")
    return Border(left=s, right=s, top=s, bottom=s)

def _c(ws, row, col, val="", bg=C_BLANC, fg=C_PREV_FG, bold=False,
       italic=False, align="center", size=10, num_fmt=None):
    from openpyxl.cell import MergedCell
    c = ws.cell(row=row, column=col)
    if isinstance(c, MergedCell):
        return c   # cellule fusionnee, on ne peut pas ecrire dedans
    c.value = val
    c.fill = _fill(bg); c.font = _font(bold=bold, color=fg, size=size, italic=italic)
    c.alignment = _align(h=align); c.border = _border()
    if num_fmt: c.number_format = num_fmt
    return c

# ─── LOGIQUE RECONCILIATION ───────────────────────────────────────
def _get_parts(series_store, annee):
    tot = series_store.get("Total")
    if tot is None: return {}
    t = float(tot[tot.index.year == annee].sum())
    if t == 0: return {}
    parts = {}
    for seg in SEG_TO_AXE:
        s = series_store.get(seg)
        parts[seg] = (float(s[s.index.year == annee].sum()) / t
                      if s is not None
                      else 1.0 / len(AXE_TO_SEGS[SEG_TO_AXE[seg]]))
    return parts

def _val_ann(forecasts, series_store, key, yr, approche, bu_axe, cache=None):
    fc = forecasts.get((key, yr), {})
    if approche == "top_down":
        if key == "Total": return fc.get("annuel", 0)
        axe = SEG_TO_AXE.get(key)
        if axe:
            td = fc.get("annuel_td", {}).get(axe)
            if td is not None: return round(float(td), 3)
        return fc.get("annuel", 0)
    segs = AXE_TO_SEGS.get(bu_axe, [])
    if key == "Total":
        return round(sum(forecasts.get((s, yr), {}).get("annuel", 0) for s in segs), 3)
    if SEG_TO_AXE.get(key) == bu_axe: return fc.get("annuel", 0)
    total_bu = sum(forecasts.get((s, yr), {}).get("annuel", 0) for s in segs)
    parts = (cache or {}).get(yr - 1) or _get_parts(series_store, yr - 1)
    return round(total_bu * parts.get(key, 0), 3)

def _val_mens(forecasts, series_store, key, yr, approche, bu_axe, cache=None):
    fc = forecasts.get((key, yr), {})
    if approche == "top_down":
        if key == "Total": return fc.get("fc", np.zeros(12))
        axe = SEG_TO_AXE.get(key)
        if axe:
            td = fc.get("top_down", {}).get(axe)
            if td is not None: return np.array(td)
        return fc.get("fc", np.zeros(12))
    segs = AXE_TO_SEGS.get(bu_axe, [])
    if key == "Total":
        arrs = [forecasts.get((s, yr), {}).get("fc", np.zeros(12)) for s in segs]
        return np.sum(arrs, axis=0) if arrs else np.zeros(12)
    if SEG_TO_AXE.get(key) == bu_axe: return fc.get("fc", np.zeros(12))
    arrs = [forecasts.get((s, yr), {}).get("fc", np.zeros(12)) for s in segs]
    bu_m  = np.sum(arrs, axis=0) if arrs else np.zeros(12)
    parts = (cache or {}).get(yr - 1) or _get_parts(series_store, yr - 1)
    return np.array([v * parts.get(key, 0) for v in bu_m])

def _get_label(approche, bu_axe):
    if approche == "top_down": return APPROCHE_LABELS["top_down"]
    return APPROCHE_LABELS.get(f"bu_{bu_axe}", approche)

def _load_escales():
    try:
        with open("forecasts_escales.pkl", "rb") as f:
            raw = pickle.load(f)
        return raw["forecasts"], raw["meta"]
    except Exception:
        return None, None

def _load_conteneurs():
    try:
        with open("forecasts_conteneurs.pkl", "rb") as f:
            raw = pickle.load(f)
        return raw["forecasts"], raw["meta"]
    except Exception:
        return None, None

# ─────────────────────────────────────────────────────────────────
# SECTION ESCALES — construction des lignes
# ─────────────────────────────────────────────────────────────────
def _section_esc_lt(ws, fc_esc, mdl_esc, annee_max_data, annees_fc,
                    row, COL_HYP):
    """Ecrit la section escales (long terme) et retourne la prochaine ligne."""
    if not fc_esc or not mdl_esc:
        return row

    ann_hist = mdl_esc["ann_total_hist"]
    yr_last  = mdl_esc["annee_fin"]
    wmape    = mdl_esc["wmape"]
    parts_yr = mdl_esc["parts_terminaux"]

    def esc_v(cle, yr):
        if yr == annee_max_data:
            if cle == "TOTAL":
                return ann_hist.get(yr, 0)
            if cle == "PORT_COM":
                # réalisé : somme exacte des terminaux
                return sum(int(fc_esc[("historique", yr)]["segments"]
                               .get(g, np.zeros(12)).sum()) for g in SEGS_ESC)
            return int(fc_esc[("historique", yr)]["segments"]
                       .get(cle, np.zeros(12)).sum())
        if cle == "TOTAL":   return fc_esc[yr]["annuel"]
        if cle == "PORT_COM":
            # prévisions : PORT_COM = TOTAL (tout le port est Port de Commerce)
            return fc_esc[yr]["annuel"]
        return int(fc_esc[yr]["segments"].get(cle, np.zeros(12)).sum())

    # En-tete section
    ws.merge_cells(f"A{row}:B{row}")
    _c(ws, row, 1, "Trafic global de navires (en nombre)",
       bg=C_GRP1, fg=C_GRP1_FG, bold=True, align="left")
    for col in range(3, COL_HYP + 1):
        _c(ws, row, col, "", bg=C_GRP1)
    ws.row_dimensions[row].height = 19
    row += 1

    def _hyp(g, p):
        return (f"Top-down : Nb total escales \u00d7 {p:.2f}%"
                f" (part {g} en {yr_last})")

    LIGNES = [
        ("TOTAL",       "Nombre total d\u2019escales",
         C_TOTAL, "000000", True,
         f"Holt amortie (double lissage amorti) \u00b7 WMAPE={wmape:.1f}%"),
        ("PORT_COM",    "Port de Commerce",
         C_GRP2,  C_GRP2_FG, True, "Somme des composantes"),
        ("TC1",         "TERMINAL A CONTENEUR (TC 1)",
         C_BLANC, C_PREV_FG, True,
         _hyp("TC1",         parts_yr[yr_last].get("TC1",0))),
        ("TC2",         "TERMINAL A CONTENEUR (TC 2)",
         C_BLANC, C_PREV_FG, True,
         _hyp("TC2",         parts_yr[yr_last].get("TC2",0))),
        ("C\u00e9r\u00e9alier", "TERMINAL C\u00c9R\u00c9ALIER",
         C_BLANC, C_PREV_FG, True,
         _hyp("C\u00e9r\u00e9alier", parts_yr[yr_last].get("C\u00e9r\u00e9alier",0))),
        ("Fruitier",    "TERMINAL FRUITIER",
         C_BLANC, C_PREV_FG, True,
         _hyp("Fruitier",    parts_yr[yr_last].get("Fruitier",0))),
        ("Min\u00e9ralier", "TERMINAL MIN\u00c9RALIER",
         C_BLANC, C_PREV_FG, True,
         _hyp("Min\u00e9ralier", parts_yr[yr_last].get("Min\u00e9ralier",0))),
        ("P\u00e9trolier",  "TERMINAL P\u00c9TROLIER",
         C_BLANC, C_PREV_FG, True,
         _hyp("P\u00e9trolier",  parts_yr[yr_last].get("P\u00e9trolier",0))),
        ("Roulier",     "TERMINAL ROULIER",
         C_BLANC, C_PREV_FG, True,
         _hyp("Roulier",     parts_yr[yr_last].get("Roulier",0))),
        ("Quai Nord",   "Quai NORD (Q1-Q5)",
         C_BLANC, C_PREV_FG, True,
         _hyp("Quai Nord",   parts_yr[yr_last].get("Quai Nord",0))),
        ("Quai Ouest",  "Quai OUEST (Q6-Q10)",
         C_BLANC, C_PREV_FG, True,
         _hyp("Quai Ouest",  parts_yr[yr_last].get("Quai Ouest",0))),
        ("Autres zones","AUTRES ZONES",
         C_BLANC, C_PREV_FG, True,
         _hyp("Autres zones",parts_yr[yr_last].get("Autres zones",0))),
    ]

    row_start = row
    for cle, label, bg, fg, bold, hyp in LIGNES:
        _c(ws, row, 1, "", bg=bg)
        _c(ws, row, 2, label, bg=bg, fg=fg, bold=bold, align="left")
        _c(ws, row, 3, int(esc_v(cle, annee_max_data)),
           bg=bg, fg=C_REEL_FG if cle == "TOTAL" else fg,
           bold=bold, num_fmt="#,##0")
        for i, yr in enumerate(annees_fc):
            _c(ws, row, 4 + i, int(esc_v(cle, yr)),
               bg=bg, fg=fg, bold=bold, num_fmt="#,##0")
        _c(ws, row, COL_HYP, hyp,
           bg=C_BLANC, fg=C_PREV_FG, italic=True, size=9, align="left")
        ws.row_dimensions[row].height = 16
        row += 1

    ws.merge_cells(f"A{row_start}:A{row-1}")
    c = ws.cell(row_start, 1)
    c.value = "Trafic global de navires\n(en nombre)"
    c.fill = _fill(C_GRP1); c.font = _font(bold=True, color=C_GRP1_FG, size=9)
    c.alignment = _align(h="center")

    # Ligne IC 95%
    ws.merge_cells(f"A{row}:B{row}")
    _c(ws, row, 1,
       f"Intervalle de confiance 95 % \u2014 Total escales (bas \u2013 haut)",
       bg=C_GRIS, fg="555555", italic=True, size=8, align="left")
    _c(ws, row, 2, "", bg=C_GRIS)
    _c(ws, row, 3, "R\u00e9alis\u00e9", bg=C_GRIS, fg="555555", italic=True, size=8)
    for i, yr in enumerate(annees_fc):
        _c(ws, row, 4 + i,
           f"{fc_esc[yr]['ic_lo']:,} \u2013 {fc_esc[yr]['ic_hi']:,}",
           bg=C_GRIS, fg="555555", italic=True, size=8)
    _c(ws, row, COL_HYP, "", bg=C_GRIS)
    ws.row_dimensions[row].height = 13
    row += 1
    return row


def _section_esc_ct(ws, fc_esc, mdl_esc, annee_max_data, annee_fc,
                    row, COL_TOT, COL_HYP):
    if not fc_esc or not mdl_esc:
        return row

    yr_last = mdl_esc["annee_fin"]
    wmape   = mdl_esc["wmape"]
    parts_yr = mdl_esc["parts_terminaux"]

    def em(cle):
        if cle == "TOTAL":   return np.array(fc_esc[annee_fc]["mensuel"], dtype=float)
        if cle == "PORT_COM":
            # PORT_COM mensuel = mensuel TOTAL (tout le port est Port de Commerce)
            return np.array(fc_esc[annee_fc]["mensuel"], dtype=float)
        return np.array(fc_esc[annee_fc]["segments"].get(cle, np.zeros(12)), dtype=float)

    # En-tete
    ws.merge_cells(f"A{row}:B{row}")
    _c(ws, row, 1, "Trafic global de navires (en nombre)",
       bg=C_GRP1, fg=C_GRP1_FG, bold=True, align="left")
    for col in range(3, COL_HYP + 1):
        _c(ws, row, col, "", bg=C_GRP1)
    ws.row_dimensions[row].height = 19
    row += 1

    def _hyp(g, p):
        return (f"Top-down : Nb total escales \u00d7 {p:.2f}%"
                f" (part {g} en {yr_last})")

    LIGNES = [
        ("TOTAL",       "Nombre total d\u2019escales",
         C_TOTAL, "000000", True,
         f"Holt amortie \u00b7 WMAPE={wmape:.1f}%"
         f" \u00b7 IC95%: {fc_esc[annee_fc]['ic_lo']:,}\u2013{fc_esc[annee_fc]['ic_hi']:,}"),
        ("PORT_COM",    "Port de Commerce",
         C_GRP2,  C_GRP2_FG, True, "Somme des composantes"),
        ("TC1",         "TERMINAL A CONTENEUR (TC 1)",
         C_BLANC, C_PREV_FG, True,
         _hyp("TC1",         parts_yr[yr_last].get("TC1",0))),
        ("TC2",         "TERMINAL A CONTENEUR (TC 2)",
         C_BLANC, C_PREV_FG, True,
         _hyp("TC2",         parts_yr[yr_last].get("TC2",0))),
        ("C\u00e9r\u00e9alier", "TERMINAL C\u00c9R\u00c9ALIER",
         C_BLANC, C_PREV_FG, True,
         _hyp("C\u00e9r\u00e9alier", parts_yr[yr_last].get("C\u00e9r\u00e9alier",0))),
        ("Fruitier",    "TERMINAL FRUITIER",
         C_BLANC, C_PREV_FG, True,
         _hyp("Fruitier",    parts_yr[yr_last].get("Fruitier",0))),
        ("Min\u00e9ralier", "TERMINAL MIN\u00c9RALIER",
         C_BLANC, C_PREV_FG, True,
         _hyp("Min\u00e9ralier", parts_yr[yr_last].get("Min\u00e9ralier",0))),
        ("P\u00e9trolier",  "TERMINAL P\u00c9TROLIER",
         C_BLANC, C_PREV_FG, True,
         _hyp("P\u00e9trolier",  parts_yr[yr_last].get("P\u00e9trolier",0))),
        ("Roulier",     "TERMINAL ROULIER",
         C_BLANC, C_PREV_FG, True,
         _hyp("Roulier",     parts_yr[yr_last].get("Roulier",0))),
        ("Quai Nord",   "Quai NORD (Q1-Q5)",
         C_BLANC, C_PREV_FG, True,
         _hyp("Quai Nord",   parts_yr[yr_last].get("Quai Nord",0))),
        ("Quai Ouest",  "Quai OUEST (Q6-Q10)",
         C_BLANC, C_PREV_FG, True,
         _hyp("Quai Ouest",  parts_yr[yr_last].get("Quai Ouest",0))),
        ("Autres zones","AUTRES ZONES",
         C_BLANC, C_PREV_FG, True,
         _hyp("Autres zones",parts_yr[yr_last].get("Autres zones",0))),
    ]

    row_start = row
    for cle, label, bg, fg, bold, hyp in LIGNES:
        v = em(cle)
        _c(ws, row, 1, "", bg=bg)
        _c(ws, row, 2, label, bg=bg, fg=fg, bold=bold, align="left")
        for m in range(12):
            _c(ws, row, 3 + m, int(v[m]),
               bg=bg, fg=fg, bold=bold, num_fmt="#,##0")
        _c(ws, row, COL_TOT, int(v.sum()),
           bg=bg, fg=fg, bold=bold, num_fmt="#,##0")
        _c(ws, row, COL_HYP, hyp,
           bg=C_BLANC, fg=C_PREV_FG, italic=True, size=9, align="left")
        ws.row_dimensions[row].height = 16
        row += 1

    ws.merge_cells(f"A{row_start}:A{row-1}")
    c = ws.cell(row_start, 1)
    c.value = "Trafic global\nde navires\n(en nombre)"
    c.fill = _fill(C_GRP1); c.font = _font(bold=True, color=C_GRP1_FG, size=9)
    c.alignment = _align(h="center")

    # IC 95%
    ws.merge_cells(f"A{row}:B{row}")
    _c(ws, row, 1, f"IC 95 % Total escales {annee_fc} (bas \u2013 haut)",
       bg=C_GRIS, fg="555555", italic=True, size=8, align="left")
    _c(ws, row, 2, "", bg=C_GRIS)
    lo = fc_esc[annee_fc]["mensuel_lo"]
    hi = fc_esc[annee_fc]["mensuel_hi"]
    for m in range(12):
        _c(ws, row, 3 + m, f"{int(lo[m])}\u2013{int(hi[m])}",
           bg=C_GRIS, fg="555555", italic=True, size=7)
    _c(ws, row, COL_TOT,
       f"{fc_esc[annee_fc]['ic_lo']:,}\u2013{fc_esc[annee_fc]['ic_hi']:,}",
       bg=C_GRIS, fg="555555", italic=True, size=7, bold=True)
    _c(ws, row, COL_HYP, "", bg=C_GRIS)
    ws.row_dimensions[row].height = 13
    row += 1
    return row




# ─────────────────────────────────────────────────────────────────
# SECTIONS CONTENEURS
# ─────────────────────────────────────────────────────────────────
_SEGS_CNT_TERM = ['TC1', 'TC2', 'Fruitier', 'Roulier', 'Autres zones']
_C_CNT_HEADER  = "1F3864"

def _cnt_v(cle, yr, fc_cnt, ann_hist, annee_max_data):
    """Valeur annuelle d'un segment conteneur."""
    if yr == annee_max_data:
        h = fc_cnt[('historique', yr)]
        if cle == 'TOTAL':      return ann_hist.get(yr, 0)
        if cle == 'TC1':        return int(h['segments_term']['TC1'].sum())
        if cle == 'TC2':        return int(h['segments_term']['TC2'].sum())
        if cle == 'Fruitier':   return int(h['segments_term']['Fruitier'].sum())
        if cle == 'Roulier':    return int(h['segments_term']['Roulier'].sum())
        if cle == 'Autres':     return int(h['segments_term']['Autres zones'].sum())
        if cle == 'Transb':
            return int(h['segments_dest']['Transb. TC2'].sum() +
                       h['segments_dest']['Transb. habituel'].sum())
        if cle == 'TransbTC2':  return int(h['segments_dest']['Transb. TC2'].sum())
        if cle == 'TransbHab':  return int(h['segments_dest']['Transb. habituel'].sum())
        if cle == 'NonTransb':  return int(h['segments_dest']['Non transb.'].sum())
    else:
        if cle == 'TOTAL':      return fc_cnt[yr]['annuel']
        if cle == 'TC1':        return int(fc_cnt[yr]['segments_term']['TC1'].sum())
        if cle == 'TC2':        return int(fc_cnt[yr]['segments_term']['TC2'].sum())
        if cle == 'Fruitier':   return int(fc_cnt[yr]['segments_term']['Fruitier'].sum())
        if cle == 'Roulier':    return int(fc_cnt[yr]['segments_term']['Roulier'].sum())
        if cle == 'Autres':     return int(fc_cnt[yr]['segments_term']['Autres zones'].sum())
        if cle == 'Transb':
            return (fc_cnt[yr]['transb_tc2_ann'] + fc_cnt[yr]['transb_hab_ann'])
        if cle == 'TransbTC2':  return fc_cnt[yr]['transb_tc2_ann']
        if cle == 'TransbHab':  return fc_cnt[yr]['transb_hab_ann']
        if cle == 'NonTransb':  return int(fc_cnt[yr]['segments_dest']['Non transb.'].sum())
    return 0


_LIGNES_CNT = [
    # (cle, label, bg, fg, bold, hypothese_template)
    # hypothese_template peut utiliser {pt} pour les parts terminaux
    # et {pd} pour les parts destination
    ('TOTAL',     "Trafic total conteneurs",
     C_TOTAL, "000000", True,
     "Holt amortie (double lissage amorti) \u00b7 err={err:.1f}% \u00b7 train 2023-N"),
    ('TC1',       "TERMINAL A CONTENEUR (TC 1)",
     C_BLANC, C_PREV_FG, True,
     "Top-down : Trafic total conteneurs \u00d7 {pt_TC1:.2f}% (part TC1 en {yr})"),
    ('TC2',       "TERMINAL A CONTENEUR (TC 2)",
     C_BLANC, C_PREV_FG, True,
     "Top-down : Trafic total conteneurs \u00d7 {pt_TC2:.2f}% (part TC2 en {yr})"),
    ('Fruitier',  "TERMINAL FRUITIER",
     C_BLANC, C_PREV_FG, True,
     "Top-down : Trafic total conteneurs \u00d7 {pt_Fruitier:.2f}% (part Fruitier en {yr})"),
    ('Roulier',   "TERMINAL ROULIER",
     C_BLANC, C_PREV_FG, True,
     "Top-down : Trafic total conteneurs \u00d7 {pt_Roulier:.2f}% (part Roulier en {yr})"),
    ('Autres',    "AUTRES ZONES",
     C_BLANC, C_PREV_FG, True,
     "Top-down : Trafic total conteneurs \u00d7 {pt_Autres:.2f}% (part Autres zones en {yr})"),
    ('Transb',    "Trafic conteneurs transbord\u00e9s",
     C_GRP2,  C_GRP2_FG, True,
     "Somme : Trafic conteneurs transbord\u00e9s TC2 + Trafic conteneurs transbord\u00e9s habituel"),
    ('TransbTC2', "Transbord\u00e9s TC2",
     C_BLANC, C_PREV_FG, True,
     "Constant = r\u00e9alis\u00e9 {yr} : {tc2:,} EVP/an"),
    ('TransbHab', "Transbord\u00e9s habituel",
     C_BLANC, C_PREV_FG, True,
     "R\u00e9siduel : Trafic total \u2212 (Non transbord\u00e9 + Transbord\u00e9 TC2)"),
    ('NonTransb', "Trafic conteneurs non transbord\u00e9s",
     C_GRP1,  C_GRP1_FG, True,
     "Top-down : Trafic total conteneurs \u00d7 {pt_NT:.2f}% (part Non transb. en {yr})"),
]

def _section_cnt_lt(ws, fc_cnt, mdl_cnt, annee_max_data, annees_fc, row, COL_HYP):
    """Ecrit la section conteneurs dans la feuille long terme."""
    if not fc_cnt or not mdl_cnt:
        return row
    ann_hist = mdl_cnt['ann_total_hist']
    yr_last  = mdl_cnt['annee_fin']
    err_tot  = mdl_cnt['err_tot']
    wmape_nt = mdl_cnt['wmape_nt']

    # En-tete section
    ws.merge_cells(f"A{row}:B{row}")
    _c(ws, row, 1, "Trafic total conteneurs (EVP)",
       bg=_C_CNT_HEADER, fg="FFFFFF", bold=True, align="left")
    for col in range(3, COL_HYP + 1):
        _c(ws, row, col, "", bg=_C_CNT_HEADER)
    ws.row_dimensions[row].height = 19
    row += 1

    row_start = row
    for cle, label, bg, fg, bold, hyp in _LIGNES_CNT:
        _c(ws, row, 1, "", bg=bg)
        _c(ws, row, 2, label, bg=bg, fg=fg, bold=bold, align="left")
        # Realise
        v_r = _cnt_v(cle, annee_max_data, fc_cnt, ann_hist, annee_max_data)
        _c(ws, row, 3, int(v_r), bg=bg,
           fg=C_REEL_FG if cle == "TOTAL" else fg,
           bold=bold, num_fmt="#,##0")
        # Previsions
        for i, yr in enumerate(annees_fc):
            v = _cnt_v(cle, yr, fc_cnt, ann_hist, annee_max_data)
            _c(ws, row, 4 + i, int(v), bg=bg, fg=fg, bold=bold, num_fmt="#,##0")
        # Formater l'hypothèse avec les vraies clés
        pt = mdl_cnt['parts_term'].get(yr_last, {})
        pd_dest = mdl_cnt['parts_dest'].get(yr_last, {})
        hyp_fmt = hyp.format(
            err=mdl_cnt['err_tot'],
            yr=yr_last,
            pt_TC1=pt.get('TC1',0),
            pt_TC2=pt.get('TC2',0),
            pt_Fruitier=pt.get('Fruitier',0),
            pt_Roulier=pt.get('Roulier',0),
            pt_Autres=pt.get('Autres zones',0),
            pt_Transb=pd_dest.get('Transbordé',0),
            pt_NT=pd_dest.get('Non transb.',0),
            tc2=mdl_cnt.get('transb_tc2_2025',0),
        ) if '{' in hyp else hyp
        _c(ws, row, COL_HYP, hyp_fmt,
           bg=C_BLANC, fg=C_PREV_FG, italic=True, size=9, align="left")
        ws.row_dimensions[row].height = 16
        row += 1

    # Fusionner colonne A
    ws.merge_cells(f"A{row_start}:A{row-1}")
    c = ws.cell(row_start, 1)
    c.value = "Trafic total\nconteneurs\n(EVP)"
    c.fill  = _fill(_C_CNT_HEADER)
    c.font  = Font(name="Arial", bold=True, color="FFFFFF", size=9)
    c.alignment = _align(h="center")

    # IC 95%
    ws.merge_cells(f"A{row}:B{row}")
    _c(ws, row, 1,
       "IC 95 % Total conteneurs (bas \u2013 haut)",
       bg=C_GRIS, fg="555555", italic=True, size=8, align="left")
    _c(ws, row, 3, "R\u00e9alis\u00e9",
       bg=C_GRIS, fg="555555", italic=True, size=8)
    for i, yr in enumerate(annees_fc):
        _c(ws, row, 4 + i,
           f"{fc_cnt[yr]['ic_lo']:,} \u2013 {fc_cnt[yr]['ic_hi']:,}",
           bg=C_GRIS, fg="555555", italic=True, size=8)
    _c(ws, row, COL_HYP, "", bg=C_GRIS)
    ws.row_dimensions[row].height = 13
    row += 1
    return row


def _section_cnt_ct(ws, fc_cnt, mdl_cnt, annee_max_data, annee_fc,
                    row, COL_TOT, COL_HYP):
    """Ecrit la section conteneurs dans la feuille court terme."""
    if not fc_cnt or not mdl_cnt:
        return row
    yr_last  = mdl_cnt['annee_fin']
    err_tot  = mdl_cnt['err_tot']
    wmape_nt = mdl_cnt['wmape_nt']
    ann_hist = mdl_cnt['ann_total_hist']
    noms_m   = mdl_cnt['noms_mois']

    # En-tete
    ws.merge_cells(f"A{row}:B{row}")
    _c(ws, row, 1, "Trafic total conteneurs (EVP)",
       bg=_C_CNT_HEADER, fg="FFFFFF", bold=True, align="left")
    for col in range(3, COL_HYP + 1):
        _c(ws, row, col, "", bg=_C_CNT_HEADER)
    ws.row_dimensions[row].height = 19
    row += 1

    # Ligne realise
    ws.merge_cells(f"A{row}:B{row}")
    _c(ws, row, 1, "R\u00c9ALIS\u00c9 " + str(annee_max_data),
       bg=C_ENTETE, fg=C_REEL_FG, bold=True, align="left")
    hist = fc_cnt[('historique', yr_last)]
    reel_m = hist['mensuel']
    for m in range(12):
        _c(ws, row, 3 + m, int(reel_m[m]),
           bg=C_ENTETE, fg=C_REEL_FG, bold=True, num_fmt="#,##0")
    _c(ws, row, COL_TOT, int(reel_m.sum()),
       bg=C_ENTETE, fg=C_REEL_FG, bold=True, num_fmt="#,##0")
    _c(ws, row, COL_HYP, "Valeurs r\u00e9alis\u00e9es",
       bg=C_BLANC, fg="555555", italic=True, size=9, align="left")
    ws.row_dimensions[row].height = 18
    row += 1

    def get_mens(cle):
        import numpy as np
        if cle == 'TOTAL':     return fc_cnt[annee_fc]['mensuel']
        if cle == 'TC1':       return fc_cnt[annee_fc]['segments_term']['TC1']
        if cle == 'TC2':       return fc_cnt[annee_fc]['segments_term']['TC2']
        if cle == 'Fruitier':  return fc_cnt[annee_fc]['segments_term']['Fruitier']
        if cle == 'Roulier':   return fc_cnt[annee_fc]['segments_term']['Roulier']
        if cle == 'Autres':    return fc_cnt[annee_fc]['segments_term']['Autres zones']
        if cle == 'Transb':
            return (fc_cnt[annee_fc]['segments_dest']['Transb. TC2'] +
                    fc_cnt[annee_fc]['segments_dest']['Transb. habituel'])
        if cle == 'TransbTC2': return fc_cnt[annee_fc]['segments_dest']['Transb. TC2']
        if cle == 'TransbHab': return fc_cnt[annee_fc]['segments_dest']['Transb. habituel']
        if cle == 'NonTransb': return fc_cnt[annee_fc]['segments_dest']['Non transb.']
        return np.zeros(12)

    row_start = row
    for cle, label, bg, fg, bold, hyp in _LIGNES_CNT:
        mens = get_mens(cle)
        _c(ws, row, 1, "", bg=bg)
        _c(ws, row, 2, label, bg=bg, fg=fg, bold=bold, align="left")
        for m in range(12):
            _c(ws, row, 3 + m, int(mens[m]),
               bg=bg, fg=fg, bold=bold, num_fmt="#,##0")
        import numpy as np
        _c(ws, row, COL_TOT, int(np.array(mens).sum()),
           bg=bg, fg=fg, bold=True, num_fmt="#,##0")
        # Formater l'hypothèse avec les vraies clés
        pt = mdl_cnt['parts_term'].get(yr_last, {})
        pd_dest = mdl_cnt['parts_dest'].get(yr_last, {})
        hyp_fmt = hyp.format(
            err=mdl_cnt['err_tot'],
            yr=yr_last,
            pt_TC1=pt.get('TC1',0),
            pt_TC2=pt.get('TC2',0),
            pt_Fruitier=pt.get('Fruitier',0),
            pt_Roulier=pt.get('Roulier',0),
            pt_Autres=pt.get('Autres zones',0),
            pt_Transb=pd_dest.get('Transbordé',0),
            pt_NT=pd_dest.get('Non transb.',0),
            tc2=mdl_cnt.get('transb_tc2_2025',0),
        ) if '{' in hyp else hyp
        _c(ws, row, COL_HYP, hyp_fmt,
           bg=C_BLANC, fg=C_PREV_FG, italic=True, size=9, align="left")
        ws.row_dimensions[row].height = 16
        row += 1

    # Fusionner colonne A
    ws.merge_cells(f"A{row_start}:A{row-1}")
    c = ws.cell(row_start, 1)
    c.value = "Trafic total\nconteneurs\n(EVP)"
    c.fill  = _fill(_C_CNT_HEADER)
    c.font  = Font(name="Arial", bold=True, color="FFFFFF", size=9)
    c.alignment = _align(h="center")

    # IC 95%
    ws.merge_cells(f"A{row}:B{row}")
    _c(ws, row, 1,
       "IC 95 % Total conteneurs " + str(annee_fc) + " (bas \u2013 haut)",
       bg=C_GRIS, fg="555555", italic=True, size=8, align="left")
    lo = fc_cnt[annee_fc]['mensuel_lo']
    hi = fc_cnt[annee_fc]['mensuel_hi']
    for m in range(12):
        _c(ws, row, 3 + m,
           str(int(lo[m])) + "\u2013" + str(int(hi[m])),
           bg=C_GRIS, fg="555555", italic=True, size=7)
    _c(ws, row, COL_TOT,
       str(fc_cnt[annee_fc]['ic_lo']) + "\u2013" + str(fc_cnt[annee_fc]['ic_hi']),
       bg=C_GRIS, fg="555555", italic=True, size=7, bold=True)
    _c(ws, row, COL_HYP, "", bg=C_GRIS)
    ws.row_dimensions[row].height = 13
    row += 1
    return row

# ─────────────────────────────────────────────────────────────────
# DEFINITION DES SECTIONS MARCHANDISES
# ─────────────────────────────────────────────────────────────────
MARCH_SECTIONS_LT = [
    {
        "titre" : "Trafic global (en tonnes)",
        "col_a" : "Trafic global\n(en tonnes)",
        "lignes": [
            ("Total",            "Trafic global",
             C_TOTAL, "000000",  True, "Holt-Winters multiplicatif \u00b7 WMAPE=3.5% \u00b7 r\u00e9f\u00e9rence Top-down"),
            ("Import",           "Importations",
             C_GRP1,  C_GRP1_FG, True, "Top-down : Trafic global × 68.76% (part Import en 2025)"),
            ("Export",           "Exportations",
             C_GRP1,  C_GRP1_FG, True, "Top-down : Trafic global × 31.24% (part Export en 2025)"),
            ("March. g\u00e9n\u00e9rales", "Marchandises g\u00e9n\u00e9rales",
             C_GRP3,  C_GRP3_FG, True, "Top-down : Trafic global × 74.02% (part March. g\u00e9n\u00e9rales en 2025)"),
            ("Prod. p\u00e9troliers",      "Produits p\u00e9troliers",
             C_GRP3,  C_GRP3_FG, True, "Top-down : Trafic global × 24.33% (part Prod. p\u00e9troliers en 2025)"),
            ("Prod. de p\u00eache",        "Produits de p\u00eache",
             C_GRP3,  C_GRP3_FG, True, "Top-down : Trafic global × 1.64% (part Prod. de p\u00eache en 2025)"),
        ],
    },
    {
        "titre" : "Trafic par destination (en tonnes)",
        "col_a" : "Trafic par\ndestination (t)",
        "lignes": [
            ("National",       "National",
             C_GRP1, C_GRP1_FG, True, "Top-down : Trafic global × 73.43% (part National en 2025)"),
            ("Transit",        "Transit (BF+Mali+Niger+Pays c\u00f4tiers)",
             C_GRP1, C_GRP1_FG, True, "Top-down : Trafic global × 8.43% (part Transit en 2025)"),
            ("Transbordement", "Transbordement",
             C_GRP1, C_GRP1_FG, True, "Top-down : Trafic global × 18.13% (part Transbordement en 2025)"),
        ],
    },
    {
        "titre" : "Trafic par conteneurisation (en tonnes)",
        "col_a" : "Par\nconteneurisation",
        "lignes": [
            ("Conteneuris\u00e9",    "Conteneuris\u00e9",
             C_GRP1, C_GRP1_FG, True, "Top-down : Trafic global × 41.90% (part Conteneuris\u00e9 en 2025)"),
            ("Non conteneuris\u00e9","Non conteneuris\u00e9",
             C_GRP1, C_GRP1_FG, True, "Top-down : Trafic global × 58.10% (part Non conteneuris\u00e9 en 2025)"),
        ],
    },
]

# ─────────────────────────────────────────────────────────────────
# TABLEAU LONG TERME
# ─────────────────────────────────────────────────────────────────
def generate_xlsx_long_terme(forecasts, series_store,
                              annee_max_data, annee_min_fc, horizon,
                              approche_key="top_down", bu_axe=None):

    annees_fc = list(range(annee_min_fc, annee_min_fc + horizon))
    label     = _get_label(approche_key, bu_axe)
    fc_esc, mdl_esc = _load_escales()

    cache = {}
    for yr in annees_fc:
        if yr - 1 not in cache:
            cache[yr - 1] = _get_parts(series_store, yr - 1)

    wb = Workbook()
    ws = wb.active
    ws.title = f"Pr\u00e9visions {annee_min_fc}-{annees_fc[-1]}"
    ws.sheet_view.showGridLines = False

    COL_HYP = 4 + len(annees_fc) - 1   # derniere colonne de données + 1
    COL_HYP = 3 + len(annees_fc) + 1   # col réalisé + cols années + col hypothèse
    NB_COLS = COL_HYP

    # Ligne 1 : titre
    ws.merge_cells(f"A1:{get_column_letter(NB_COLS)}1")
    _c(ws, 1, 1, f"TRAFICS PR\u00c9VISIONNELS PORT D\u2019ABIDJAN \u2014 {annee_min_fc}\u2013{annees_fc[-1]}",
       bg=C_ENTETE, fg=C_TITRE, bold=True, size=12)
    ws.row_dimensions[1].height = 28

    # Ligne 2 : sous-titre
    ws.merge_cells(f"A2:{get_column_letter(NB_COLS)}2")
    _c(ws, 2, 1, f"R\u00e9conciliation : {label}   |   Base arr\u00eat\u00e9e au 31/12/{annee_max_data}",
       bg=C_ENTETE, fg="555555", italic=True, size=9, align="left")
    ws.row_dimensions[2].height = 16

    # Ligne 3 : en-tetes colonnes
    ws.merge_cells("A3:B3")
    _c(ws, 3, 1, "Indicateurs / Composantes",
       bg=C_ENTETE, fg=C_TITRE, bold=True)
    _c(ws, 3, 3, f"R\u00e9alis\u00e9\n{annee_max_data}",
       bg=C_ENTETE, fg=C_REEL_FG, bold=True)
    for i, yr in enumerate(annees_fc):
        _c(ws, 3, 4 + i, f"Pr\u00e9vu\n{yr}",
           bg=C_ENTETE, fg=C_TITRE, bold=True)
    _c(ws, 3, COL_HYP, "Hypoth\u00e8se de pr\u00e9diction",
       bg=C_ENTETE, fg=C_TITRE, bold=True)
    ws.row_dimensions[3].height = 34

    row = 4

    # Section escales
    row = _section_esc_lt(ws, fc_esc, mdl_esc, annee_max_data,
                          annees_fc, row, COL_HYP)

    # Section conteneurs
    fc_cnt_lt, mdl_cnt_lt = _load_conteneurs()
    row = _section_cnt_lt(ws, fc_cnt_lt, mdl_cnt_lt, annee_max_data,
                          annees_fc, row, COL_HYP)

    # Sections marchandises
    for section in MARCH_SECTIONS_LT:
        ws.merge_cells(f"A{row}:B{row}")
        _c(ws, row, 1, section["titre"],
           bg=C_GRP1, fg=C_GRP1_FG, bold=True, align="left")
        for col in range(3, NB_COLS + 1):
            _c(ws, row, col, "", bg=C_GRP1)
        ws.row_dimensions[row].height = 19
        row += 1

        row_start = row
        for key, label_l, bg, fg, bold, hyp in section["lignes"]:
            serie = series_store.get(key)
            reel  = (float(serie[serie.index.year == annee_max_data].sum())
                     if serie is not None else 0)
            _c(ws, row, 1, "", bg=bg)
            _c(ws, row, 2, label_l, bg=bg, fg=fg, bold=bold, align="left")
            _c(ws, row, 3, round(reel, 3), bg=bg,
               fg=C_REEL_FG if key == "Total" else fg,
               bold=bold, num_fmt="#,##0.000")
            for i, yr in enumerate(annees_fc):
                v = _val_ann(forecasts, series_store, key, yr,
                             approche_key, bu_axe, cache)
                _c(ws, row, 4 + i, round(float(v), 3),
                   bg=bg, fg=fg, bold=bold, num_fmt="#,##0.000")
            _c(ws, row, COL_HYP, hyp,
               bg=C_BLANC, fg=C_PREV_FG, italic=True, size=9, align="left")
            ws.row_dimensions[row].height = 16
            row += 1

        ws.merge_cells(f"A{row_start}:A{row-1}")
        c = ws.cell(row_start, 1)
        c.value = section["col_a"]
        c.fill = _fill(C_GRP1); c.font = _font(bold=True, color=C_GRP1_FG, size=9)
        c.alignment = _align(h="center")

    # Ligne croissance trafic global
    ws.merge_cells(f"A{row}:B{row}")
    _c(ws, row, 1, "Taux de croissance annuel \u2014 Trafic global (Mt)",
       bg=C_GRIS, fg="555555", bold=True, size=9, align="left")
    _c(ws, row, 2, "", bg=C_GRIS)
    _c(ws, row, 3, "\u2014", bg=C_GRIS, fg="555555", size=9)
    serie_tot = series_store.get("Total")
    prev = float(serie_tot[serie_tot.index.year == annee_max_data].sum()) if serie_tot is not None else 0
    for i, yr in enumerate(annees_fc):
        v = _val_ann(forecasts, series_store, "Total", yr,
                     approche_key, bu_axe, cache)
        pct = (float(v) / prev - 1) if prev else 0
        _c(ws, row, 4 + i, round(pct * 100, 1),
           bg="E8F5E9" if pct >= 0 else "FFEBEE",
           fg="196B24" if pct >= 0 else "C00000",
           bold=True, size=9, num_fmt='0.0"%"')
        prev = float(v)
    _c(ws, row, COL_HYP, "", bg=C_GRIS)
    ws.row_dimensions[row].height = 16

    # Dimensionnement
    ws.column_dimensions["A"].width = 22
    ws.column_dimensions["B"].width = 36
    ws.column_dimensions["C"].width = 16
    for i in range(len(annees_fc)):
        ws.column_dimensions[get_column_letter(4 + i)].width = 14
    ws.column_dimensions[get_column_letter(COL_HYP)].width = 62
    ws.freeze_panes = "C4"

    buf = io.BytesIO()
    wb.save(buf); buf.seek(0)
    return buf


# ─────────────────────────────────────────────────────────────────
# TABLEAU COURT TERME
# ─────────────────────────────────────────────────────────────────
def generate_xlsx_court_terme(forecasts, series_store,
                               annee_max_data, annee_fc,
                               approche_key="top_down", bu_axe=None):

    label    = _get_label(approche_key, bu_axe)
    fc_esc, mdl_esc = _load_escales()
    cache    = {annee_fc - 1: _get_parts(series_store, annee_fc - 1)}

    COL_TOT = 15   # colonne TOTAL ANNUEL
    COL_HYP = 16   # colonne Hypothese
    NB_COLS = 16

    wb = Workbook()
    ws = wb.active
    ws.title = f"Pr\u00e9visions mensuelles {annee_fc}"
    ws.sheet_view.showGridLines = False

    # Ligne 1
    ws.merge_cells(f"A1:{get_column_letter(NB_COLS)}1")
    _c(ws, 1, 1, f"PR\u00c9VISIONS MENSUELLES \u2014 PORT D\u2019ABIDJAN {annee_fc}",
       bg=C_ENTETE, fg=C_TITRE, bold=True, size=12)
    ws.row_dimensions[1].height = 28

    # Ligne 2
    ws.merge_cells(f"A2:{get_column_letter(NB_COLS)}2")
    _c(ws, 2, 1, f"R\u00e9conciliation : {label}   |   Base arr\u00eat\u00e9e au 31/12/{annee_max_data}",
       bg=C_ENTETE, fg="555555", italic=True, size=9, align="left")
    ws.row_dimensions[2].height = 16

    # Ligne 3 en-tetes
    ws.merge_cells("A3:B3")
    _c(ws, 3, 1, "Indicateurs / Composantes",
       bg=C_ENTETE, fg=C_TITRE, bold=True)
    for m, nom in enumerate(MOIS_NOMS):
        _c(ws, 3, 3 + m, nom, bg=C_ENTETE, fg=C_TITRE, bold=True, size=9)
    _c(ws, 3, COL_TOT, "TOTAL\nANNUEL",
       bg=C_ENTETE, fg=C_TITRE, bold=True)
    _c(ws, 3, COL_HYP, "Hypoth\u00e8se de pr\u00e9diction",
       bg=C_ENTETE, fg=C_TITRE, bold=True)
    ws.row_dimensions[3].height = 34

    row = 4

    # ── Ligne réalisé ESCALES 2025 ──────────────────────────────
    ws.merge_cells(f"A{row}:B{row}")
    _c(ws, row, 1, f"R\u00c9ALIS\u00c9 {annee_max_data} \u2014 Escales (nombre)",
       bg=C_ENTETE, fg=C_REEL_FG, bold=True, align="left")
    # Réalisé escales par mois depuis forecasts_escales
    if fc_esc and mdl_esc:
        yr_last_esc = mdl_esc["annee_fin"]
        reel_esc = fc_esc[("historique", yr_last_esc)]["mensuel"]
        tot_esc  = int(fc_esc[("historique", yr_last_esc)]["annuel"])
    else:
        reel_esc = [0]*12; tot_esc = 0
    for m in range(12):
        _c(ws, row, 3 + m, int(reel_esc[m]),
           bg=C_ENTETE, fg=C_REEL_FG, bold=True, num_fmt="#,##0")
    _c(ws, row, COL_TOT, tot_esc,
       bg=C_ENTETE, fg=C_REEL_FG, bold=True, num_fmt="#,##0")
    _c(ws, row, COL_HYP, "Valeurs r\u00e9alis\u00e9es",
       bg=C_BLANC, fg="555555", italic=True, size=9, align="left")
    ws.row_dimensions[row].height = 18
    row += 1

    # Section escales (prévisions 2026)
    row = _section_esc_ct(ws, fc_esc, mdl_esc, annee_max_data,
                          annee_fc, row, COL_TOT, COL_HYP)


    # Section conteneurs
    fc_cnt_ct, mdl_cnt_ct = _load_conteneurs()
    row = _section_cnt_ct(ws, fc_cnt_ct, mdl_cnt_ct, annee_max_data,
                          annee_fc, row, COL_TOT, COL_HYP)
    # ── Ligne réalisé MARCHANDISES 2025 ─────────────────────────
    ws.merge_cells(f"A{row}:B{row}")
    _c(ws, row, 1, f"R\u00c9ALIS\u00c9 {annee_max_data} \u2014 Marchandises (Mt)",
       bg=C_ENTETE, fg=C_REEL_FG, bold=True, align="left")
    serie_tot = series_store.get("Total")
    reel_mois = []
    if serie_tot is not None:
        ra = serie_tot[serie_tot.index.year == annee_max_data]
        reel_mois = [round(float(ra[ra.index.month == m].sum()), 3)
                     for m in range(1, 13)]
    for m in range(12):
        _c(ws, row, 3 + m, reel_mois[m] if reel_mois else 0,
           bg=C_ENTETE, fg=C_REEL_FG, bold=True, num_fmt="#,##0.000")
    _c(ws, row, COL_TOT, round(sum(reel_mois), 3) if reel_mois else 0,
       bg=C_ENTETE, fg=C_REEL_FG, bold=True, num_fmt="#,##0.000")
    _c(ws, row, COL_HYP, "Valeurs r\u00e9alis\u00e9es",
       bg=C_BLANC, fg="555555", italic=True, size=9, align="left")
    ws.row_dimensions[row].height = 18
    row += 1

    # Sections marchandises
    MARCH_CT = [
        {
            "titre" : "Trafic global (en tonnes)",
            "col_a" : "Trafic global\n(en tonnes)",
            "lignes": [
                ("Total",            "Trafic global (Mt)",
                 C_TOTAL, "000000",  True, "Holt-Winters multiplicatif \u00b7 WMAPE=3.5% \u00b7 r\u00e9f\u00e9rence Top-down"),
                ("Import",           "Importations",
                 C_GRP1,  C_GRP1_FG, True, "Top-down : Trafic global × 68.76% (part Import en 2025)"),
                ("Export",           "Exportations",
                 C_GRP1,  C_GRP1_FG, True, "Top-down : Trafic global × 31.24% (part Export en 2025)"),
                ("March. g\u00e9n\u00e9rales", "Marchandises g\u00e9n\u00e9rales",
                 C_GRP3,  C_GRP3_FG, True, "Top-down : Trafic global × 74.02% (part March. g\u00e9n\u00e9rales en 2025)"),
                ("Prod. p\u00e9troliers",      "Produits p\u00e9troliers",
                 C_GRP3,  C_GRP3_FG, True, "Top-down : Trafic global × 24.33% (part Prod. p\u00e9troliers en 2025)"),
                ("Prod. de p\u00eache",        "Produits de p\u00eache",
                 C_GRP3,  C_GRP3_FG, True, "Top-down : Trafic global × 1.64% (part Prod. de p\u00eache en 2025)"),
            ],
        },
        {
            "titre" : "Trafic par destination (en tonnes)",
            "col_a" : "Trafic par\ndestination (t)",
            "lignes": [
                ("National",       "National",
                 C_GRP1, C_GRP1_FG, True, "Top-down : Trafic global × 73.43% (part National en 2025)"),
                ("Transit",        "Transit",
                 C_GRP1, C_GRP1_FG, True, "Top-down : Trafic global × 8.43% (part Transit en 2025)"),
                ("Transbordement", "Transbordement",
                 C_GRP1, C_GRP1_FG, True, "Top-down : Trafic global × 18.13% (part Transbordement en 2025)"),
            ],
        },
        {
            "titre" : "Trafic par conteneurisation (en tonnes)",
            "col_a" : "Par\nconteneurisation",
            "lignes": [
                ("Conteneuris\u00e9",    "Conteneuris\u00e9",
                 C_GRP1, C_GRP1_FG, True, "Top-down : Trafic global × 41.90% (part Conteneuris\u00e9 en 2025)"),
                ("Non conteneuris\u00e9","Non conteneuris\u00e9",
                 C_GRP1, C_GRP1_FG, True, "Top-down : Trafic global × 58.10% (part Non conteneuris\u00e9 en 2025)"),
            ],
        },
    ]

    for section in MARCH_CT:
        ws.merge_cells(f"A{row}:B{row}")
        _c(ws, row, 1, section["titre"],
           bg=C_GRP1, fg=C_GRP1_FG, bold=True, align="left")
        for col in range(3, NB_COLS + 1):
            _c(ws, row, col, "", bg=C_GRP1)
        ws.row_dimensions[row].height = 19
        row += 1

        row_start = row
        for key, label_l, bg, fg, bold, hyp in section["lignes"]:
            fc_m = _val_mens(forecasts, series_store, key, annee_fc,
                             approche_key, bu_axe, cache)
            fc_m = [round(float(v), 3) for v in fc_m]
            _c(ws, row, 1, "", bg=bg)
            _c(ws, row, 2, label_l, bg=bg, fg=fg, bold=bold, align="left")
            for m in range(12):
                _c(ws, row, 3 + m, fc_m[m],
                   bg=bg, fg=fg, bold=bold, num_fmt="#,##0.000")
            _c(ws, row, COL_TOT, round(sum(fc_m), 3),
               bg=bg, fg=fg, bold=True, num_fmt="#,##0.000")
            _c(ws, row, COL_HYP, hyp,
               bg=C_BLANC, fg=C_PREV_FG, italic=True, size=9, align="left")
            ws.row_dimensions[row].height = 16
            row += 1

        ws.merge_cells(f"A{row_start}:A{row-1}")
        c = ws.cell(row_start, 1)
        c.value = section["col_a"]
        c.fill = _fill(C_GRP1)
        c.font = _font(bold=True, color=C_GRP1_FG, size=9)
        c.alignment = _align(h="center")

    # Dimensionnement
    ws.column_dimensions["A"].width = 20
    ws.column_dimensions["B"].width = 34
    for m in range(12):
        ws.column_dimensions[get_column_letter(3 + m)].width = 10
    ws.column_dimensions[get_column_letter(COL_TOT)].width = 13
    ws.column_dimensions[get_column_letter(COL_HYP)].width = 60
    ws.freeze_panes = "C5"

    buf = io.BytesIO()
    wb.save(buf); buf.seek(0)
    return buf
